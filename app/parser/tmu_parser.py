"""
tmu_parser.py
=============
Parse TMU (Trường Đại học Thương Mại) timetable XLSX files and convert
them into a list of AcademicEvent domain objects ready for PostgreSQL
ingestion and Google Calendar sync.

Architecture note
-----------------
- start_time / end_time are real ``datetime.time`` values derived from
  TMU period numbers (50-minute slots, see PERIOD_MAP).
- fingerprint uses start_time (HH:MM) instead of start_period so it
  remains meaningful even if the period schedule changes.
- All fields map 1-to-1 to the target DB schema; no post-processing needed.

Usage
-----
    from tmu_parser import parse_events

    events = parse_events("TKB_HK2_…xlsx")               # relative path OK
    events = parse_events("TKB_HK2_…xlsx", verbose=True)
    for ev in events:
        print(ev.to_dict())
"""

from __future__ import annotations

import os
import re
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# TMU period → wall-clock start time
# ---------------------------------------------------------------------------

PERIOD_MAP: Dict[int, time] = {
    1:  time(7,  0),
    2:  time(7,  50),
    3:  time(8,  40),
    4:  time(9,  30),
    5:  time(10, 20),
    6:  time(11, 10),
    7:  time(12, 30),
    8:  time(13, 20),
    9:  time(14, 10),
    10: time(15,  0),
    11: time(15, 50),
    12: time(16, 40),
}

PERIOD_DURATION_MINUTES = 50  # each TMU period is 50 minutes


def period_to_time(period: int) -> time:
    """Return the wall-clock start time for a TMU period number (1–12).

    Raises KeyError if *period* is not in PERIOD_MAP.
    """
    if period not in PERIOD_MAP:
        raise KeyError(
            f"Period {period!r} is not a valid TMU period. "
            f"Expected one of {sorted(PERIOD_MAP)}."
        )
    return PERIOD_MAP[period]


def _compute_end_time(start: time, duration_periods: int) -> time:
    """end_time = start_time + duration * 50 minutes."""
    total_minutes = duration_periods * PERIOD_DURATION_MINUTES
    dt = datetime.combine(date.today(), start) + timedelta(minutes=total_minutes)
    return dt.time()


# ---------------------------------------------------------------------------
# Domain model
# ---------------------------------------------------------------------------

@dataclass
class AcademicEvent:
    """One teaching slot extracted from the timetable.

    weekday uses the Vietnamese sheet convention:
        THỨ 2 → 2  (Monday)
        THỨ 3 → 3  (Tuesday)
        …
        THỨ 7 → 7  (Saturday)
        CHỦ NHẬT → 7  (Sunday — distinguish via start_date weekday if needed)

    start_time / end_time are Python ``datetime.time`` objects.
    PostgreSQL column type: TIME WITHOUT TIME ZONE.
    """

    course_code: str
    course_name: str
    instructor:  str              # Placeholder — not in the raw file
    weekday:     int              # 2–7
    start_time:  time
    end_time:    time
    room:        str
    start_date:  Optional[date]
    end_date:    Optional[date]
    duration:    int              # number of 50-min periods
    fingerprint: str = field(init=False)

    def __post_init__(self) -> None:
        hhmm = self.start_time.strftime("%H:%M")
        self.fingerprint = f"{self.course_code}-{self.weekday}-{hhmm}-{self.room}"

    def to_dict(self) -> dict:
        """Serialize to a plain dict suitable for psycopg2 / SQLAlchemy execute()."""
        return {
            "course_code": self.course_code,
            "course_name": self.course_name,
            "instructor":  self.instructor,
            "weekday":     self.weekday,
            "start_time":  self.start_time.strftime("%H:%M"),
            "end_time":    self.end_time.strftime("%H:%M"),
            "room":        self.room,
            "start_date":  self.start_date.isoformat() if self.start_date else None,
            "end_date":    self.end_date.isoformat()   if self.end_date   else None,
            "duration":    self.duration,
            "fingerprint": self.fingerprint,
        }

    def __repr__(self) -> str:
        return (
            f"AcademicEvent({self.course_code!r}, weekday={self.weekday}, "
            f"{self.start_time.strftime('%H:%M')}–{self.end_time.strftime('%H:%M')}, "
            f"room={self.room!r}, dates={self.start_date}→{self.end_date})"
        )


# ---------------------------------------------------------------------------
# Vietnamese weekday label → weekday int
# ---------------------------------------------------------------------------

_WEEKDAY_LABEL_MAP: Dict[str, int] = {
    "THỨ 2":    2,
    "THỨ 3":    3,
    "THỨ 4":    4,
    "THỨ 5":    5,
    "THỨ 6":    6,
    "THỨ 7":    7,
    "CHỦ NHẬT": 7,
}


def _normalise(s) -> str:
    return re.sub(r"\s+", " ", str(s).strip().upper())


def _normalise_room(value) -> str:
    """Strip whitespace; unify all online variants to 'ONLINE'."""
    if value is None:
        return ""
    s = str(value).strip()
    if s.upper() in ("ONLINE", "TRỰC TUYẾN", "TRUC TUYEN"):
        return "ONLINE"
    return s


# ---------------------------------------------------------------------------
# Step 1 – detect header row
# ---------------------------------------------------------------------------

def detect_header_row(ws: Worksheet, search_limit: int = 20) -> int:
    """Return 1-based row index containing 'THỨ 2'.

    Raises ValueError when not found within *search_limit* rows.
    """
    for row_idx in range(1, search_limit + 1):
        for cell in ws[row_idx]:
            if cell.value and _normalise(str(cell.value)) == "THỨ 2":
                return row_idx
    raise ValueError(
        f"Could not find weekday header ('THỨ 2') in the first {search_limit} rows."
    )


# ---------------------------------------------------------------------------
# Step 2 – build column map
# ---------------------------------------------------------------------------

@dataclass
class _ColumnMap:
    course_code_col: int
    course_name_col: int
    start_date_col:  int
    end_date_col:    int
    # (weekday_int, start_period_col, duration_col, room_col) — all 0-based
    weekday_groups: List[Tuple[int, int, int, int]] = field(default_factory=list)


def build_column_map(ws: Worksheet, header_row: int) -> _ColumnMap:
    """Derive column positions from the three header rows without hardcoding.

    Row layout (relative to header_row, 1-based):
      +0 : main labels  (STT | MÃ LỚP | TÊN HP | … | THỨ 2 | … | NGÀY BẮT ĐẦU …)
      +1 : secondary    (Trực tiếp | Trực tuyến for teaching-mode cols)
      +2 : sub-headers  (Tiết bắt đầu | Số tiết | Phòng  — repeated per weekday)
    """
    main_row = list(
        ws.iter_rows(min_row=header_row,     max_row=header_row,     values_only=True)
    )[0]
    sub2_row = list(
        ws.iter_rows(min_row=header_row + 2, max_row=header_row + 2, values_only=True)
    )[0]

    course_code_col: Optional[int] = None
    course_name_col: Optional[int] = None
    start_date_col:  Optional[int] = None
    end_date_col:    Optional[int] = None
    weekday_anchors: List[Tuple[int, int]] = []

    for col_idx, val in enumerate(main_row):
        if val is None:
            continue
        label = _normalise(str(val))

        if label in ("MÃ LỚP HỌC PHẦN", "MÃ LHP"):
            course_code_col = col_idx
        elif label in ("TÊN HỌC PHẦN", "TÊN HP"):
            course_name_col = col_idx
        elif "BẮT ĐẦU" in label:
            start_date_col = col_idx
        elif "KẾT THÚC" in label:
            end_date_col = col_idx
        elif label in _WEEKDAY_LABEL_MAP:
            weekday_anchors.append((_WEEKDAY_LABEL_MAP[label], col_idx))

    weekday_groups: List[Tuple[int, int, int, int]] = []
    for weekday_int, anchor_col in weekday_anchors:
        sp_col = dur_col = room_col = None
        for offset in range(3):
            c = anchor_col + offset
            if c >= len(sub2_row):
                break
            v = sub2_row[c]
            if v is None:
                continue
            norm = _normalise(str(v))
            if "BẮT ĐẦU" in norm:
                sp_col = c
            elif "SỐ TIẾT" in norm:
                dur_col = c
            elif "PHÒNG" in norm:
                room_col = c

        if sp_col is not None and dur_col is not None and room_col is not None:
            weekday_groups.append((weekday_int, sp_col, dur_col, room_col))
        else:
            # Fallback: three consecutive columns
            weekday_groups.append(
                (weekday_int, anchor_col, anchor_col + 1, anchor_col + 2)
            )

    missing = [
        name for name, val in [
            ("MÃ LỚP HỌC PHẦN", course_code_col),
            ("TÊN HỌC PHẦN",    course_name_col),
            ("NGÀY BẮT ĐẦU",    start_date_col),
            ("NGÀY KẾT THÚC",   end_date_col),
        ] if val is None
    ]
    if missing:
        raise ValueError(f"Could not locate required columns: {missing}")
    if not weekday_groups:
        raise ValueError("No weekday columns detected in header row.")

    return _ColumnMap(
        course_code_col=course_code_col,
        course_name_col=course_name_col,
        start_date_col=start_date_col,
        end_date_col=end_date_col,
        weekday_groups=weekday_groups,
    )


# ---------------------------------------------------------------------------
# Private helpers
# ---------------------------------------------------------------------------

def _parse_date(value) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _safe_int(value) -> Optional[int]:
    if value is None:
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        return None


def _is_footer(row_values) -> bool:
    for val in row_values:
        if val and "ghi chú" in str(val).lower():
            return True
    return False


def _resolve_path(file_path: str) -> str:
    """Return an absolute path.

    Relative paths are resolved from os.getcwd() so the parser works
    correctly regardless of where the calling script lives.
    """
    return file_path if os.path.isabs(file_path) else os.path.abspath(file_path)


# ---------------------------------------------------------------------------
# Step 3 – parse events (main public API)
# ---------------------------------------------------------------------------

def parse_events(
    file_path: str,
    sheet_name: Optional[str] = None,
    verbose: bool = False,
) -> List[AcademicEvent]:
    """Parse a TMU timetable XLSX file → ``List[AcademicEvent]``.

    Parameters
    ----------
    file_path:
        Path to the .xlsx file. Relative paths are resolved via
        ``os.path.abspath(file_path)``.
    sheet_name:
        Worksheet name. Uses the first sheet when omitted.
    verbose:
        Print progress and skip/count statistics to stdout.

    Returns
    -------
    List[AcademicEvent]
        One entry per (row × weekday) slot with a valid start_period.
    """
    abs_path = _resolve_path(file_path)

    if verbose:
        print(f"[tmu_parser] Loading : {abs_path}")

    wb = openpyxl.load_workbook(abs_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]

    if verbose:
        print(f"[tmu_parser] Sheet   : '{ws.title}'  "
              f"({ws.max_row} rows × {ws.max_column} cols)")

    header_row = detect_header_row(ws)
    col_map    = build_column_map(ws, header_row)
    data_start = header_row + 3

    if verbose:
        print(f"[tmu_parser] Header row : {header_row}  |  "
              f"Data starts at row : {data_start}")
        print(f"[tmu_parser] Weekday groups : {len(col_map.weekday_groups)}")

    events: List[AcademicEvent] = []
    n_empty = n_footer = n_no_code = n_bad_period = n_bad_weekday = 0

    for row in ws.iter_rows(min_row=data_start, values_only=True):

        if all(v is None for v in row):
            n_empty += 1
            continue

        if _is_footer(row):
            n_footer += 1
            continue

        course_code = row[col_map.course_code_col]
        course_name = row[col_map.course_name_col]

        if not course_code:
            n_no_code += 1
            continue

        course_code = str(course_code).strip()
        course_name = str(course_name).strip() if course_name else ""
        start_date  = _parse_date(row[col_map.start_date_col])
        end_date    = _parse_date(row[col_map.end_date_col])

        for weekday_int, sp_col, dur_col, room_col in col_map.weekday_groups:
            start_period = _safe_int(row[sp_col])
            if start_period is None:
                continue   # no class on this weekday for this row

            if not (1 <= weekday_int <= 7):
                n_bad_weekday += 1
                if verbose:
                    print(f"[tmu_parser] SKIP bad weekday={weekday_int} ({course_code})")
                continue

            if start_period not in PERIOD_MAP:
                n_bad_period += 1
                if verbose:
                    print(f"[tmu_parser] SKIP unknown period={start_period} ({course_code})")
                continue

            duration   = _safe_int(row[dur_col]) or 1
            room       = _normalise_room(row[room_col])
            start_time = period_to_time(start_period)
            end_time   = _compute_end_time(start_time, duration)

            events.append(AcademicEvent(
                course_code = course_code,
                course_name = course_name,
                instructor  = "",
                weekday     = weekday_int,
                start_time  = start_time,
                end_time    = end_time,
                room        = room,
                start_date  = start_date,
                end_date    = end_date,
                duration    = duration,
            ))

    if verbose:
        print(
            f"\n[tmu_parser] ── Summary ─────────────────────────────────\n"
            f"  Events generated      : {len(events)}\n"
            f"  Skipped (empty rows)  : {n_empty}\n"
            f"  Skipped (footer rows) : {n_footer}\n"
            f"  Skipped (no code)     : {n_no_code}\n"
            f"  Skipped (bad period)  : {n_bad_period}\n"
            f"  Skipped (bad weekday) : {n_bad_weekday}\n"
            f"[tmu_parser] ─────────────────────────────────────────────"
        )

    return events


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import json
    import sys

    cli_path    = sys.argv[1] if len(sys.argv) > 1 else "timetable.xlsx"
    cli_verbose = "--verbose" in sys.argv or "-v" in sys.argv

    results = parse_events(cli_path, verbose=cli_verbose)

    if not cli_verbose:
        print(f"Parsed {len(results)} events.")

    print("\nFirst 5 events:")
    for ev in results[:5]:
        print(" ", ev)

    print("\nSample JSON (first event):")
    if results:
        print(json.dumps(results[0].to_dict(), ensure_ascii=False, indent=2))
